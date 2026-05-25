"""사육소 계산 출력 엑셀 생성기.

양식 (사용자 제공 '사육소 계산기(2026부터사용)' 참고):

  시트1) YYYY년 송아지개체  — 어린소(매입월령 ≤ 13) 출생연도별 리스트
    - 컬럼: NO. / 개체식별번호 / 성별 / 출생일자 / 매입일 / 매입월령 / 월별 개월령
    - 데이터 아래에 COUNTIF / 카테고리 통계 표들 (엑셀 수식)
    - 우측 하단: 총사육두수 / 월평균두수 / 비과세 / 기준초과두수

  시트2) YYYY년 기준 소 목록  — 모든 소 (전체 농장 합산)
    - 컬럼: NO. / 개체식별번호 / 성별 / 출생일자 / 매입일 / 매입월령 / 상태 / 현재월령 / 비고
"""
from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import date, datetime
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="FFE699")
END_STATUSES = {"도축", "폐사", "양수도"}

STATUS_DISPLAY = {
    "사육": "전산등록",
    "도축": "도축출하",
    "폐사": "폐사",
    "양수도": "양수",
}


@dataclass
class CowRow:
    cattle_no: str
    sex: str | None
    birth_date: str | None
    acquisition_date: str | None
    status: str
    status_date: str | None


# ────────────────────────────────────────────────────────────
# 헬퍼
# ────────────────────────────────────────────────────────────
def _parse_ymd(s: str | None) -> date | None:
    if not s:
        return None
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def _format_cattle_no(no: str) -> str:
    n = "".join(ch for ch in no if ch.isdigit())
    if len(n) == 12:
        return f"{n[:3]} {n[3:7]} {n[7:11]} {n[11]}"
    return no


def _short_no(no: str) -> str:
    """비고용 끝 4자리 (12자리 식별번호의 8~11번째)."""
    n = "".join(ch for ch in no if ch.isdigit())
    if len(n) == 12:
        return n[7:11]
    return n[-4:]


def _months_diff(a: date, b: date) -> int:
    """a → b 까지의 월 차이 +1 (a 가 그 월에 출생/매입한 시점 = 1)."""
    return (b.year - a.year) * 12 + (b.month - a.month) + 1


def _generate_month_columns(start: date, end: date) -> list[tuple[int, int, str]]:
    """(year, month, label) 리스트. start ~ end 까지 매월."""
    result = []
    y, m = start.year, start.month
    while (y, m) <= (end.year, end.month):
        label = f"{y % 100}년 {m}월"
        result.append((y, m, label))
        m += 1
        if m == 13:
            m = 1
            y += 1
    return result


def _set_header(cell, text: str) -> None:
    cell.value = text
    cell.font = Font(bold=True)
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")


# ────────────────────────────────────────────────────────────
# 시트1) 송아지개체 + 통계 표들
# ────────────────────────────────────────────────────────────
def _build_calf_sheets(
    wb: Workbook,
    cows: list[CowRow],
    month_cols: list[tuple[int, int, str]],
) -> None:
    by_year: dict[int, list[CowRow]] = {}
    for c in cows:
        birth = _parse_ymd(c.birth_date)
        acq = _parse_ymd(c.acquisition_date)
        if not birth or not acq:
            continue
        if _months_diff(birth, acq) > 13:
            continue  # 성축 매입은 제외
        by_year.setdefault(birth.year, []).append(c)

    for year in sorted(by_year.keys(), reverse=True):
        sheet_name = f"{year}년 송아지개체"
        ws = wb.create_sheet(title=sheet_name[:31])
        _write_calf_sheet(ws, by_year[year], month_cols)


def _write_calf_sheet(ws, rows: list[CowRow], month_cols: list[tuple[int, int, str]]) -> None:
    base_headers = ["NO.", "개체식별번호", "성별", "출생일자", "매입일", "매입월령"]
    month_labels = [lbl for _, _, lbl in month_cols]
    all_headers = base_headers + month_labels

    for i, h in enumerate(all_headers, start=1):
        _set_header(ws.cell(row=1, column=i), h)

    sorted_rows = sorted(
        rows, key=lambda c: (_parse_ymd(c.acquisition_date) or date.min, c.cattle_no),
        reverse=True,
    )

    for i, c in enumerate(sorted_rows, start=1):
        birth = _parse_ymd(c.birth_date)
        acq = _parse_ymd(c.acquisition_date)
        acq_mo = _months_diff(birth, acq)
        end_dt = _parse_ymd(c.status_date) if c.status in END_STATUSES else None

        row_values = [
            i,
            _format_cattle_no(c.cattle_no),
            c.sex or "",
            birth,
            acq,
            acq_mo,
        ]
        for y, m, _ in month_cols:
            cell_month = date(y, m, 1)
            diff = (y - acq.year) * 12 + (m - acq.month)
            is_acquired = diff >= 0
            is_alive = (
                end_dt is None
                or cell_month < date(end_dt.year, end_dt.month, 1)
                or (end_dt.year == y and end_dt.month == m)
            )
            if is_acquired and is_alive:
                row_values.append(acq_mo + diff)
            else:
                row_values.append(None)
        ws.append(row_values)

    # 컬럼 폭
    widths = [5, 18, 6, 12, 12, 9]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i in range(len(base_headers) + 1, len(all_headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 8
    ws.freeze_panes = "G2"

    # 통계 영역
    n_cows = len(sorted_rows)
    if n_cows == 0:
        return
    data_last_row = 1 + n_cows  # 1=header row, then data
    n_months = len(month_cols)
    first_month_col = 7  # G
    last_month_col = first_month_col + n_months - 1
    total_col = last_month_col + 1

    # ── COUNTIF 자동 합계 (비어있지 않은 셀 개수)
    countif_row = data_last_row + 2
    for c_idx in range(first_month_col, last_month_col + 1):
        col_letter = get_column_letter(c_idx)
        ws.cell(
            row=countif_row, column=c_idx,
            value=f'=COUNTIF({col_letter}2:{col_letter}{data_last_row},"<>")',
        )

    # ── 표1: 6-13개월 / 5이하 어린소 / 합계
    t1_header_row = countif_row + 3
    _set_header(ws.cell(row=t1_header_row, column=first_month_col - 1), "")
    for i, (_, m, _) in enumerate(month_cols):
        _set_header(ws.cell(row=t1_header_row, column=first_month_col + i), f"{m}월")
    _set_header(ws.cell(row=t1_header_row, column=total_col), "합계")

    r6_13 = t1_header_row + 1
    r5_below = t1_header_row + 2
    r_all = t1_header_row + 3
    ws.cell(row=r6_13, column=first_month_col - 1, value="6-13개월육성우").font = Font(bold=True)
    ws.cell(row=r5_below, column=first_month_col - 1, value="5개월이하어린소").font = Font(bold=True)
    ws.cell(row=r_all, column=first_month_col - 1, value="1-13개월전체합계").font = Font(bold=True)
    ws.cell(row=r_all, column=first_month_col - 1).fill = HEADER_FILL

    for i in range(n_months):
        col = first_month_col + i
        col_letter = get_column_letter(col)
        rng = f'{col_letter}2:{col_letter}{data_last_row}'
        ws.cell(row=r6_13, column=col, value=f'=COUNTIFS({rng},">=6",{rng},"<=13")')
        ws.cell(row=r5_below, column=col, value=f'=COUNTIFS({rng},"<=5")')
        ws.cell(
            row=r_all, column=col,
            value=f"={get_column_letter(col)}{r6_13}+{get_column_letter(col)}{r5_below}",
        ).fill = HEADER_FILL

    # 합계 셀
    first_letter = get_column_letter(first_month_col)
    last_letter = get_column_letter(last_month_col)
    for r_idx in (r6_13, r5_below, r_all):
        ws.cell(
            row=r_idx, column=total_col,
            value=f"=SUMPRODUCT(--({first_letter}{r_idx}:{last_letter}{r_idx}>0),"
                  f"{first_letter}{r_idx}:{last_letter}{r_idx})",
        )
    ws.cell(row=r_all, column=total_col).fill = HEADER_FILL

    # ── 표2: 월사육두수 (사용자 수동 입력)
    t2_header_row = r_all + 3
    _set_header(ws.cell(row=t2_header_row, column=first_month_col - 1), "구 분")
    for i, (_, m, _) in enumerate(month_cols):
        _set_header(ws.cell(row=t2_header_row, column=first_month_col + i), f"{m:02d}월")
    _set_header(ws.cell(row=t2_header_row, column=total_col), "합계")

    r_total = t2_header_row + 1
    ws.cell(row=r_total, column=first_month_col - 1, value="월사육두수").font = Font(bold=True)
    ws.cell(row=r_total, column=first_month_col - 1).fill = HEADER_FILL
    for i in range(n_months):
        ws.cell(row=r_total, column=first_month_col + i, value=0)
    ws.cell(
        row=r_total, column=total_col,
        value=f"=SUMPRODUCT(--({first_letter}{r_total}:{last_letter}{r_total}>0),"
              f"{first_letter}{r_total}:{last_letter}{r_total})",
    )

    # ── 표3: 사육두수 분석 (육성우 0.5마리 절사)
    t3_header_row = r_total + 3
    _set_header(ws.cell(row=t3_header_row, column=first_month_col - 1), "(육성우 0.5마리 절사)")
    for i, (_, m, _) in enumerate(month_cols):
        _set_header(ws.cell(row=t3_header_row, column=first_month_col + i), f"{m}월")
    _set_header(ws.cell(row=t3_header_row, column=total_col), "합계")

    r_under5 = t3_header_row + 1
    r_14plus = t3_header_row + 2
    r_grow_half = t3_header_row + 3
    r_grand = t3_header_row + 4
    r_over50 = t3_header_row + 5

    ws.cell(row=r_under5, column=first_month_col - 1, value="월사육두수-5개월이하").font = Font(bold=True)
    ws.cell(row=r_14plus, column=first_month_col - 1, value="14개월이상성축").font = Font(bold=True)
    ws.cell(row=r_grow_half, column=first_month_col - 1, value="6-13개월육성우(1/2)").font = Font(bold=True)
    ws.cell(row=r_grand, column=first_month_col - 1, value="총 사육두수 합계").font = Font(bold=True)
    ws.cell(row=r_grand, column=first_month_col - 1).fill = HEADER_FILL
    ws.cell(row=r_over50, column=first_month_col - 1, value="매월 50두 초과두수").font = Font(bold=True)

    for i in range(n_months):
        col = first_month_col + i
        L = get_column_letter(col)
        ws.cell(row=r_under5, column=col, value=f"={L}{r_total}-{L}{r_5_below if False else r5_below}")
        ws.cell(row=r_14plus, column=col, value=f"={L}{r_under5}-{L}{r6_13}")
        ws.cell(row=r_grow_half, column=col, value=f"=ROUNDDOWN({L}{r6_13}/2,0)")
        ws.cell(row=r_grand, column=col,
                value=f"={L}{r_grow_half}+{L}{r_14plus}").fill = HEADER_FILL
        ws.cell(row=r_over50, column=col, value=f"=MAX(0,{L}{r_grand}-50)")

    for r_idx in (r_under5, r_14plus, r_grow_half, r_grand, r_over50):
        ws.cell(
            row=r_idx, column=total_col,
            value=f"=SUMPRODUCT(--({first_letter}{r_idx}:{last_letter}{r_idx}>0),"
                  f"{first_letter}{r_idx}:{last_letter}{r_idx})",
        )
    ws.cell(row=r_grand, column=total_col).fill = HEADER_FILL

    # ── 우측 하단 요약
    summary_col_label = total_col - 1
    summary_col_value = total_col
    r_total_all = r_over50 + 3
    ws.cell(row=r_total_all, column=summary_col_label, value="총사육두수").font = Font(bold=True)
    ws.cell(row=r_total_all, column=summary_col_value,
            value=f"={get_column_letter(total_col)}{r_grand}")
    ws.cell(row=r_total_all + 1, column=summary_col_label, value="월평균두수").font = Font(bold=True)
    ws.cell(
        row=r_total_all + 1, column=summary_col_value,
        value=f'=IF(COUNTIF({first_letter}{r_grand}:{last_letter}{r_grand},">0")=0,0,'
              f'ROUND(SUMPRODUCT(--({first_letter}{r_grand}:{last_letter}{r_grand}>0),'
              f'{first_letter}{r_grand}:{last_letter}{r_grand})/'
              f'COUNTIF({first_letter}{r_grand}:{last_letter}{r_grand},">0"),0))',
    )
    ws.cell(row=r_total_all + 2, column=summary_col_label, value="비과세").font = Font(bold=True)
    ws.cell(row=r_total_all + 2, column=summary_col_value, value=50)
    ws.cell(row=r_total_all + 3, column=summary_col_label, value="기준초과두수").font = Font(bold=True)
    ws.cell(row=r_total_all + 3, column=summary_col_label).fill = HEADER_FILL
    ws.cell(
        row=r_total_all + 3, column=summary_col_value,
        value=f"={get_column_letter(summary_col_value)}{r_total_all}-"
              f"{get_column_letter(summary_col_value)}{r_total_all + 2}",
    ).fill = HEADER_FILL


# ────────────────────────────────────────────────────────────
# 시트2) 기준 소 목록
# ────────────────────────────────────────────────────────────
def _build_base_list_sheet(
    wb: Workbook,
    cows: list[CowRow],
    *,
    reference_date: date,
    year: int,
) -> None:
    ws = wb.create_sheet(title=f"{year}년 기준 소 목록"[:31])
    headers = ["NO.", "개체식별번호", "성별", "출생일자", "매입일", "매입월령",
               "상태", "현재월령", "비고"]
    for i, h in enumerate(headers, start=1):
        _set_header(ws.cell(row=1, column=i), h)

    valid = [c for c in cows if _parse_ymd(c.birth_date) and _parse_ymd(c.acquisition_date)]
    valid.sort(
        key=lambda c: (_parse_ymd(c.acquisition_date), c.cattle_no),
        reverse=True,
    )

    for i, c in enumerate(valid, start=1):
        birth = _parse_ymd(c.birth_date)
        acq = _parse_ymd(c.acquisition_date)
        acq_mo = _months_diff(birth, acq)
        cur_mo = _months_diff(birth, reference_date)
        status_text = STATUS_DISPLAY.get(c.status, c.status)
        sdate = _parse_ymd(c.status_date)
        if sdate and c.status in END_STATUSES and c.status != "양수도":
            status_text = f"{status_text}({sdate.isoformat()})"

        ws.append([
            i,
            _format_cattle_no(c.cattle_no),
            c.sex or "",
            birth,
            acq,
            acq_mo,
            status_text,
            cur_mo,
            _short_no(c.cattle_no),
        ])

    widths = [5, 18, 6, 12, 12, 9, 22, 9, 9]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


# ────────────────────────────────────────────────────────────
# 메인 빌더
# ────────────────────────────────────────────────────────────
def build_workbook(
    cows: Iterable[CowRow],
    *,
    doc_date_range: tuple[date, date],
) -> bytes:
    start, end = doc_date_range
    month_cols = _generate_month_columns(start, end)
    cows_list = list(cows)

    wb = Workbook()
    wb.remove(wb.active)

    _build_calf_sheets(wb, cows_list, month_cols)
    _build_base_list_sheet(wb, cows_list, reference_date=end, year=end.year)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
