"""안심농장 사육수 계산 메인 파이프라인.

순서:
  1) input/ 폴더의 엑셀 파일 수집 (날짜순)
  2) 각 파일마다:
     - 이미 처리된 (doc_name, farm_name) 조합이면 스킵
     - 농장 정보, 개체 리스트 파싱해 DB upsert
     - 같은 농장의 이전 처리 doc과 비교해 "사라진 개체" 추출
     - 각 개체를 API로 조회해 status/status_date 업데이트
     - processed_docs 마킹
  3) 출력 엑셀에 현재 연도 시트 생성 및 활성 개체 작성
"""
from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

import db
from api_client import fetch_status
from excel_parser import parse_farm_excel, parse_filename

PROJECT_ROOT = Path(__file__).parent
INPUT_DIR = PROJECT_ROOT / "input"
OUTPUT_DIR = PROJECT_ROOT / "output"
OUTPUT_FILE = OUTPUT_DIR / "사육소계산(현재사용중).xlsx"


def collect_input_files() -> list[Path]:
    files = []
    for p in sorted(INPUT_DIR.glob("*.xls")):
        meta = parse_filename(p)
        if meta is None:
            print(f"  [skip] 파일명 형식 불일치: {p.name}")
            continue
        files.append(p)
    # doc_date 오름차순 정렬
    files.sort(key=lambda p: parse_filename(p).doc_date)
    return files


def get_cattle_nos_in_doc(farm_name: str, doc_date: str) -> set[str]:
    rows = db.get_cattle_by_farm_doc_date(farm_name, doc_date)
    return {r["cattle_no"] for r in rows}


def process_file(path: Path, *, skip_api: bool = False) -> None:
    meta = parse_filename(path)
    if meta is None:
        return

    if db.is_doc_processed(meta.doc_name, meta.farm_name):
        print(f"[skip] 이미 처리됨: {meta.doc_name}")
        return

    print(f"\n[처리] {meta.doc_name} (농장: {meta.farm_name}, 기준일: {meta.doc_date})")

    farm_info, cattle_rows = parse_farm_excel(path)
    farm_name = farm_info["farm_name"] or meta.farm_name
    db.upsert_farm(farm_name, farm_info["farm_id"], farm_info["owner"], farm_info["address"])

    for r in cattle_rows:
        r["farm_name"] = farm_name

    prev_doc_date = db.get_prev_doc_date(farm_name, meta.doc_date)
    inserted, updated = db.upsert_cattle_batch(cattle_rows, meta.doc_date)
    print(f"  개체: 신규 {inserted}, 갱신 {updated} (총 {len(cattle_rows)})")

    if prev_doc_date:
        prev_set = get_cattle_nos_in_doc(farm_name, prev_doc_date)
        curr_set = {r["cattle_no"] for r in cattle_rows}
        missing = prev_set - curr_set
        print(f"  전월({prev_doc_date}) 대비 사라진 개체: {len(missing)}")
        for no in sorted(missing):
            if skip_api:
                print(f"    [dry-run] {no}")
                continue
            info = fetch_status(no)
            db.update_cattle_status(no, info.status, info.status_date)
            print(f"    {no} → {info.status} ({info.status_date or '-'})")
    else:
        print("  (이전 처리 기록 없음 — 비교 생략)")

    db.mark_doc_processed(meta.doc_name, meta.farm_name, meta.doc_date)


def write_output_excel() -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)
    year = datetime.now().year
    sheet_name = f"{year}년"

    if OUTPUT_FILE.exists():
        from openpyxl import load_workbook
        wb = load_workbook(OUTPUT_FILE)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
    else:
        wb = Workbook()
        # 첫 기본 시트 제거
        default = wb.active
        wb.remove(default)

    ws = wb.create_sheet(title=sheet_name)
    headers = ["농장명", "개체식별번호", "소의종류", "성별", "출생일자", "모 개체번호",
               "최초확인일", "최종확인일", "상태"]
    ws.append(headers)

    rows = db.get_active_cattle_all()
    for r in rows:
        ws.append([
            r["farm_name"],
            r["cattle_no"],
            r["breed"],
            r["sex"],
            r["birth_date"],
            r["mother_no"],
            r["first_seen_doc_date"],
            r["last_seen_doc_date"],
            r["status"],
        ])

    widths = [10, 18, 10, 6, 12, 18, 12, 12, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    wb.save(OUTPUT_FILE)
    print(f"\n[저장] {OUTPUT_FILE} (시트: {sheet_name}, {len(rows)}건)")


def main(skip_api: bool = False) -> None:
    db.init_db()
    files = collect_input_files()
    print(f"input 파일 {len(files)}개")
    for p in files:
        process_file(p, skip_api=skip_api)
    write_output_excel()
    print("\n완료.")


if __name__ == "__main__":
    skip_api = "--skip-api" in sys.argv
    main(skip_api=skip_api)
